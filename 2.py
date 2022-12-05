from openpyxl import load_workbook
import pandas as pd

data = {'Name': ['Tom', 'Joseph', 'Krish', 'John'], 'Age': [20, 21, 19, 18]}
df = pd.DataFrame(data)

filename = "excels2.xlsx"
workbook = load_workbook(filename)
writer = pd.ExcelWriter(filename, engine='openpyxl')
writer.book = workbook
writer.sheets = {ws.title: ws for ws in workbook.worksheets}

df.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, index = False, header= False)

writer.close()